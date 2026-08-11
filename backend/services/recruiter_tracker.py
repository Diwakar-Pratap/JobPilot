"""
Recruiter Tracker Service
Extracts contact info (emails, phones) from job descriptions and persists
them into an Excel workbook for easy CRM-like access.
"""
import os
import re
from datetime import datetime
from typing import List, Tuple

from openpyxl import Workbook, load_workbook

EXPORTS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "exports")
EXCEL_PATH = os.path.join(EXPORTS_DIR, "recruiters.xlsx")

HEADERS = [
    "Name", "Email", "Phone", "Company",
    "Job Title", "Job URL", "Source", "Date Found",
]

# ── regex helpers ──────────────────────────────────────────────

EMAIL_RE = re.compile(
    r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}",
)

PHONE_RE = re.compile(
    r"(?:\+?\d{1,3}[\s\-.]?)?"       # optional country code
    r"(?:\(?\d{2,4}\)?[\s\-.]?)?"     # optional area code
    r"\d{3,4}[\s\-.]?\d{3,4}",        # main number
)


def _extract_emails(text: str) -> List[str]:
    """Return deduplicated list of email addresses found in *text*."""
    # Filter out common image / asset false positives
    ignore_suffixes = (".png", ".jpg", ".jpeg", ".gif", ".svg", ".webp")
    return list({
        m for m in EMAIL_RE.findall(text)
        if not any(m.lower().endswith(s) for s in ignore_suffixes)
    })


def _extract_phones(text: str) -> List[str]:
    """Return deduplicated list of phone-number strings found in *text*."""
    raw = PHONE_RE.findall(text)
    # Keep only candidates with ≥ 7 digits (skip short number sequences)
    phones: List[str] = []
    for p in raw:
        digits = re.sub(r"\D", "", p)
        if len(digits) >= 7:
            phones.append(p.strip())
    return list(set(phones))


def _ensure_workbook() -> Workbook:
    """Open or create the recruiter workbook with headers."""
    os.makedirs(EXPORTS_DIR, exist_ok=True)
    if os.path.exists(EXCEL_PATH):
        return load_workbook(EXCEL_PATH)
    wb = Workbook()
    ws = wb.active
    ws.title = "Recruiters"
    ws.append(HEADERS)
    wb.save(EXCEL_PATH)
    return wb


def _existing_emails(ws) -> set:
    """Return the set of emails already recorded (column B, 1-indexed col 2)."""
    emails: set = set()
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        if row[1]:
            emails.add(row[1].lower().strip())
    return emails


# ── public API ─────────────────────────────────────────────────

def extract_and_save(
    description: str,
    company: str,
    job_title: str,
    job_url: str,
) -> int:
    """
    Parse *description* for emails & phones, append new contacts to the
    recruiter Excel sheet, and return the number of **new** rows added.
    Duplicates (by email) are silently skipped.
    """
    emails = _extract_emails(description)
    phones = _extract_phones(description)

    if not emails and not phones:
        return 0

    wb = _ensure_workbook()
    ws = wb.active
    known = _existing_emails(ws)
    added = 0
    today = datetime.now().strftime("%Y-%m-%d")

    # Pair each email with the first available phone (or blank)
    contacts: List[Tuple[str, str]] = []
    for i, email in enumerate(emails):
        phone = phones[i] if i < len(phones) else ""
        contacts.append((email, phone))

    # If there are leftover phones with no email, still record them
    if len(phones) > len(emails):
        for phone in phones[len(emails):]:
            contacts.append(("", phone))

    for email, phone in contacts:
        if email and email.lower().strip() in known:
            continue
        ws.append([
            "",          # Name — unknown
            email,
            phone,
            company,
            job_title,
            job_url,
            "auto-extract",
            today,
        ])
        if email:
            known.add(email.lower().strip())
        added += 1

    if added:
        wb.save(EXCEL_PATH)

    return added
